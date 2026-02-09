from email import message
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

from users.serializer import UserProfileSerializer


from .forms import *


# def login_page(request):
#     forms = LoginForm()
#     if request.method == 'POST':
#         forms = LoginForm(request.POST)
#         if forms.is_valid():
#             username = forms.cleaned_data['username']
#             password = forms.cleaned_data['password']
#             print(username)
#             print(password)
#             user = authenticate(username=username, password=password)
#             if user:
#                 login(request, user)

#                 if user.is_doctor:
#                     print('---------------------------------')
#                     print('---------------------------------')
#                     print('---------------------------------')
#                 return redirect('dashboard')
#             else:
#                 messages.error(request, 'wrong username password')
#     context = {'form': forms}
#     return render(request, 'adminLogin.html', context)

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.generics import ListAPIView
from rest_framework_simplejwt.tokens import RefreshToken
from .models import User, OTP, UserToken  # Your custom user model
from .serializer import NotificationSerializer
from .otp_utils import create_and_send_otp, verify_otp
from customer.models import Notification


class SendOTPView(APIView):
    """Send OTP to mobile number"""
    
    def post(self, request):
        mobile = request.data.get("mobile")
        
        if not mobile:
            return Response({"error": "Mobile number is required"}, status=400)
        
        # Clean mobile number
        mobile = ''.join(filter(str.isdigit, str(mobile)))
        
        if len(mobile) < 10:
            return Response({"error": "Invalid mobile number"}, status=400)
        
        # Create and send OTP
        otp_obj, success, message = create_and_send_otp(mobile)
        
        if success:
            return Response({
                "message": "OTP sent successfully",
                "mobile": mobile
            }, status=200)
        else:
            return Response({"error": message}, status=400)


class VerifyOTPView(APIView):
    """Verify OTP and return JWT tokens"""
    
    def post(self, request):
        mobile = request.data.get("mobile")
        otp_code = request.data.get("otp")
        
        if not mobile or not otp_code:
            return Response({"error": "Mobile number and OTP are required"}, status=400)
        
        # Clean mobile number
        mobile = ''.join(filter(str.isdigit, str(mobile)))
        
        # Verify OTP
        otp_obj, is_valid, message = verify_otp(mobile, otp_code)
        
        if not is_valid:
            return Response({"error": message}, status=400)
        
        # Get or create user
        user = User.objects.filter(mobile=mobile).first()
        created = False
        
        if not user:
            user = User.objects.create(
                mobile=mobile,
                is_active=True
            )
            created = True
        
        # Generate JWT tokens
        refresh = RefreshToken.for_user(user)
        
        return Response({
            "access": str(refresh.access_token),
            "refresh": str(refresh),
            "user": {
                "id": user.id,
                "mobile": user.mobile,
                "created": created
            }
        }, status=200)


class SignupView(APIView):
    """Signup user after OTP verification"""
    
    def post(self, request):
        mobile = request.data.get("mobile")
        otp_code = request.data.get("otp")
        user_type = request.data.get("user_type")
        city = request.data.get("city")
        area = request.data.get("area")
        name = request.data.get("name")
        email = request.data.get("email")

        if not mobile or not otp_code or not user_type or not city or not area:
            return Response({"error": "Mobile, OTP, user_type, city, and area are required"}, status=400)

        # Clean mobile number
        mobile = ''.join(filter(str.isdigit, str(mobile)))
        
        # Verify OTP first
        otp_obj, is_valid, message = verify_otp(mobile, otp_code)
        
        if not is_valid:
            return Response({"error": message}, status=400)

        try:
            # Role flags
            role_flags = {
                "is_customer": False,
                "is_doctor": False,
            }

            if f"is_{user_type}" not in role_flags:
                return Response({"error": "Invalid user_type"}, status=400)

            user = User.objects.filter(mobile=mobile, city=city, area=area).first()
            created = False

            if user:
                # Already exists – check role
                existing_roles = [key for key, value in {
                    "customer": user.is_customer,
                    "doctor": user.is_doctor,
                }.items() if value]

                if existing_roles and user_type not in existing_roles:
                    return Response({
                        "error": f"This number is already registered as a {existing_roles[0]}. Cannot register again as {user_type}."
                    }, status=400)

            else:
                role_flags[f"is_{user_type}"] = True

                # Ensure email is unique
                if email and User.objects.filter(email=email).exists():
                    return Response({"error": "This email is already in use."}, status=400)

                user = User.objects.create(
                    mobile=mobile,
                    first_name=name or "",
                    email=email or "",
                    city_id=city,
                    area_id=area,
                    **role_flags
                )
                created = True

            refresh = RefreshToken.for_user(user)
            return Response({
                "access": str(refresh.access_token),
                "refresh": str(refresh),
                "user": {
                    "id": user.id,
                    "mobile": user.mobile,
                    "email": user.email,
                    "name": user.first_name,
                    "city": user.city.name if user.city else "",
                    "area": user.area.name if user.area else "",
                    "user_type": user_type,
                    "created": created
                }
            })

        except Exception as e:
            return Response({"error": str(e)}, status=400)

class RegisterDeviceTokenAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        token = (request.data.get("token") or "").strip()
        if not token:
            return Response({"error": "Token required"}, status=400)
        UserToken.objects.update_or_create(
            user=request.user, token=token, defaults={}
        )
        return Response({"success": True})


class NotificationListAPIView(ListAPIView):
    """List all notifications for the logged-in user (works for both doctor and customer)."""
    permission_classes = [IsAuthenticated]
    serializer_class = NotificationSerializer

    def get_queryset(self):
        return Notification.objects.filter(user=self.request.user).order_by("-created_at")


class MarkNotificationReadAPIView(APIView):
    """Mark a single notification as read. PATCH or POST."""
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk=None):
        pk = pk or self.kwargs.get("pk")
        if pk is None:
            return Response({"error": "Notification ID required."}, status=status.HTTP_400_BAD_REQUEST)
        notification = Notification.objects.filter(user=request.user, pk=pk).first()
        if not notification:
            return Response({"error": "Notification not found."}, status=status.HTTP_404_NOT_FOUND)
        notification.is_read = True
        notification.save()
        return Response({"success": True, "id": notification.id})


from customer.models import customer

class LoginAPIView(APIView):
    """Login user after OTP verification"""
    
    def post(self, request):
        mobile = request.data.get("mobile")
        otp_code = request.data.get("otp")
        user_type = request.data.get("user_type")

        if not mobile or not otp_code:
            return Response({"error": "Mobile number and OTP are required"}, status=400)

        # Clean mobile number
        mobile = ''.join(filter(str.isdigit, str(mobile)))
        
        # Verify OTP
        otp_obj, is_valid, message = verify_otp(mobile, otp_code)
        
        if not is_valid:
            return Response({"error": message}, status=400)

        try:
            user = User.objects.filter(mobile=mobile).first()
            created = False

            if user:
                if not user.is_active:
                    return Response(
                        {"error": "Your account has been deactivated. Please contact support."},
                        status=status.HTTP_403_FORBIDDEN
                    )
                # Existing user: type must match account — doctor+doctor and customer+customer only
                user_type_lower = (user_type or "").lower()
                # Doctor account but they sent type customer → reject
                if user.is_doctor and user_type_lower != "doctor":
                    return Response(
                        {"error": "This number is registered as a doctor. Please use doctor login."},
                        status=status.HTTP_403_FORBIDDEN
                    )
                # Customer account but they sent type doctor → reject
                if user.is_customer and user_type_lower == "doctor":
                    return Response(
                        {"error": "This number is registered as a customer. Please use customer login."},
                        status=status.HTTP_403_FORBIDDEN
                    )
            else:
                # New user: only create as customer (no doctor signup here)
                user = User.objects.create(
                    mobile=mobile,
                    is_active=True,
                    is_customer=True
                )
                customer.objects.create(user=user, is_active=True)
                created = True

            # Token creation
            refresh = RefreshToken.for_user(user)
            user_details = UserProfileSerializer(user).data

            # Calculate subscription status (for doctors)
            is_subscribed = False
            if user.is_doctor:
                is_subscribed = user.subscription_is_active

            return Response({
                "access": str(refresh.access_token),
                "refresh": str(refresh),
                "user": {
                    "id": user.id,
                    "mobile": user.mobile,
                    "role": "doctor" if user.is_doctor else "customer"
                },
                "created": created,
                "is_subscribed": is_subscribed,
                "user_details": user_details
            }, status=201 if created else 200)

        except Exception as e:
            print(f"Login failed: {e}")
            return Response({"error": str(e)}, status=400)





from .permissions import *


class UsergetView(APIView):
    permission_classes = [IsCustomer]

    def get(self, request):
        user = request.user
        return Response({
            "name": user.first_name,
            "email": user.email,
        })

class UserUpdateView(APIView):
    permission_classes = [IsCustomer]

    def put(self, request):
        user = request.user
        name = request.data.get("name")
        email = request.data.get("email")

        updated = False

        if name:
            user.first_name = name
            updated = True

        if email:
            user.email = email
            updated = True

        if updated:
            user.save()
            return Response({"message": "Profile updated successfully."})
        else:
            return Response({"message": "No changes provided."}, status=status.HTTP_400_BAD_REQUEST)


class ResetPasswordView(APIView):
    """Reset password using OTP verification"""
    
    def post(self, request):
        mobile = request.data.get("mobile")
        otp_code = request.data.get("otp")
        new_password = request.data.get("new_password")

        if not mobile or not otp_code or not new_password:
            return Response({"error": "Mobile, OTP, and new_password are required"}, status=400)

        # Clean mobile number
        mobile = ''.join(filter(str.isdigit, str(mobile)))
        
        # Verify OTP
        otp_obj, is_valid, message = verify_otp(mobile, otp_code)
        
        if not is_valid:
            return Response({"error": message}, status=400)

        try:
            user = User.objects.filter(mobile=mobile).first()
            
            if not user:
                return Response({"error": "User not found"}, status=404)

            # Update password
            user.set_password(new_password)
            user.save()

            return Response({"message": "Password updated successfully."})

        except Exception as e:
            return Response({"error": str(e)}, status=400)
        



def  login_admin(request):

    forms = LoginForm()
    if request.method == 'POST':
        forms = LoginForm(request.POST)
        if forms.is_valid():
            mobile = forms.cleaned_data['mobile']
            password = forms.cleaned_data['password']
            print(mobile)
            print(password)
            user = authenticate(mobile=mobile, password=password)
            if user:
                login(request, user)

                if user.is_superuser:
                    print('---------------------------------')
                    print('---------------------------------')
                    print('---------------------------------')
                return redirect('dashboard')
            else:
                messages.error(request, 'wrong username password')
    context = {'form': forms}
    return render(request, 'adminLogin.html', context)


# def resgister_page(request):

#     forms = registerForm()
#     if request.method == 'POST':
#         forms = registerForm(request.POST)
#         if forms.is_valid():
#             forms.save()
#             username = forms.cleaned_data['username']
#             password = forms.cleaned_data['password1']
#             user = authenticate(username=username, password=password)
#             if user:
                
#                 messages.error(request, 'user already exsist')
#                 return redirect('dashboard')
#             else:
#                 return redirect('resgister')
#         else:
#             print(forms.errors)
#     else:
#         print(forms.as_p)

#         context = {'form': forms}

#         return render(request, 'users/resgister.html', context)


def logout_page(request):
    logout(request)
    return redirect('login_admin')

def user_list(request):

    data = User.objects.filter(is_doctor = True)

    return render(request, 'user_list.html', { 'data' : data})


def customer_list(request):
    """
    List all customers (users with is_customer=True)
    """
    data = User.objects.filter(is_customer=True).order_by('-date_joined')
    count = data.count()
    
    return render(request, 'customer_list.html', {'data': data, 'count': count})


@login_required(login_url='login_admin')
def export_customer_list_excel(request):
    """Export customer list to Excel"""
    data = User.objects.filter(is_customer=True).order_by('-date_joined')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    
    # Headers
    headers = ['#', 'User Email', 'First Name', 'Last Name', 'Mobile', 'Gender', 'Date of Birth', 'Address', 'Date Joined']
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for idx, user in enumerate(data, start=1):
        ws.append([
            idx,
            user.email or "-",
            user.first_name or "-",
            user.last_name or "-",
            user.mobile or "-",
            user.gender or "-",
            user.dob.strftime("%d-%m-%Y") if user.dob else "-",
            user.address or "-",
            user.date_joined.strftime("%d-%m-%Y %H:%M") if user.date_joined else "-"
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"customers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required(login_url='login_admin')
def view_doctor_details(request, doctor_id):
    """
    View all details of a doctor
    """
    from doctor.models import doctor
    
    doctor_instance = get_object_or_404(doctor, id=doctor_id)
    user = doctor_instance.user
    
    context = {
        'doctor': doctor_instance,
        'user': user,
    }
    
    return render(request, 'view_doctor.html', context)


from doctor.models import doctor

def dentist_list(request):

    data = doctor.objects.all()
    count = data.count()

    return render(request, 'list_doctor.html', { 'data' : data, 'count': count})


@login_required(login_url='login_admin')
def export_dentist_list_excel(request):
    """Export dentist list to Excel with subscription valid till date"""
    data = doctor.objects.select_related('user').all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Dentists"
    
    # Headers - including subscription_valid_to
    headers = ['#', 'User Email', 'Name', 'Mobile', 'Gender', 'Address', 'Subscription Valid From', 'Subscription Valid Till']
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for idx, doc in enumerate(data, start=1):
        user = doc.user
        full_name = f"{user.first_name or ''} {user.last_name or ''}".strip() or "-"
        address = user.address or "-"
        
        ws.append([
            idx,
            user.email or "-",
            full_name,
            user.mobile or "-",
            user.gender or "-",
            address,
            user.subscription_valid_from.strftime("%d-%m-%Y") if user.subscription_valid_from else "-",
            user.subscription_valid_to.strftime("%d-%m-%Y") if user.subscription_valid_to else "-"
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"dentists_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required(login_url='login_admin')
def update_user_subscription(request, user_id):
    """
    View and update user subscription details (for doctors)
    """
    user = get_object_or_404(User, id=user_id)
    doctor_profile = None
    
    # Get doctor profile if user is a doctor
    if user.is_doctor:
        try:
            doctor_profile = doctor.objects.get(user=user)
        except doctor.DoesNotExist:
            doctor_profile = None
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        # Handle user account activate/deactivate
        if action == 'activate_user':
            user.is_active = True
            user.save()
            messages.success(request, 'User account activated successfully!')
            return redirect('update_user_subscription', user_id=user_id)
        
        elif action == 'deactivate_user':
            user.is_active = False
            user.save()
            messages.success(request, 'User account deactivated successfully!')
            return redirect('update_user_subscription', user_id=user_id)

        elif action == 'deactivate_subscription':
            from django.utils import timezone
            yesterday = timezone.now().date() - timezone.timedelta(days=1)
            user.subscription_valid_to = yesterday
            user.save()
            messages.success(request, 'Subscription deactivated successfully (valid to set to yesterday).')
            return redirect('update_user_subscription', user_id=user_id)

        # Regular update of subscription fields
        subscription_valid_from = request.POST.get('subscription_valid_from')
        subscription_valid_to = request.POST.get('subscription_valid_to')
        subscription_received_amount = request.POST.get('subscription_received_amount')
        
        if subscription_valid_from:
            user.subscription_valid_from = subscription_valid_from
        else:
            user.subscription_valid_from = None
            
        if subscription_valid_to:
            user.subscription_valid_to = subscription_valid_to
        else:
            user.subscription_valid_to = None
        
        if subscription_received_amount:
            try:
                user.subscription_received_amount = float(subscription_received_amount)
            except (ValueError, TypeError):
                pass
        else:
            user.subscription_received_amount = 0.00
            
        user.save()
        
        messages.success(request, 'Subscription details updated successfully!')
        return redirect('update_user_subscription', user_id=user_id)
    
    context = {
        'user': user,
        'doctor': doctor_profile,
    }
    
    return render(request, 'update_user_subscription.html', context)


@login_required(login_url='login_admin')
def subscription_payment_history(request, user_id):
    """
    View payment history for user subscription
    """
    user = get_object_or_404(User, id=user_id)
    doctor_profile = None
    
    # Get doctor profile if user is a doctor
    if user.is_doctor:
        try:
            doctor_profile = doctor.objects.get(user=user)
        except doctor.DoesNotExist:
            doctor_profile = None
    
    # Get payment history (PaidDoubt payments for this user)
    from customer.models import PaidDoubt
    payment_history = PaidDoubt.objects.filter(user=user, payment_status='paid').order_by('-created_at')
    
    # Calculate total received
    total_received = sum(payment.amount for payment in payment_history)
    
    context = {
        'user': user,
        'doctor': doctor_profile,
        'payment_history': payment_history,
        'total_received': total_received,
    }
    
    return render(request, 'subscription_payment_history.html', context)




from rest_framework import viewsets, mixins
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.decorators import action


class UserProfileViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    @action(detail=False, methods=['get', 'put'], url_path='me')
    def me(self, request):
        user = request.user

        if request.method == 'GET':
            serializer = UserProfileSerializer(user)
            return Response(serializer.data)

        elif request.method == 'PUT':

            serializer = UserProfileSerializer(user, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)

            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        





from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.db import transaction, IntegrityError


class DeleteAccountAPIView(APIView):
    """Delete the authenticated user's account. Class-based for consistency with other DRF endpoints."""
    permission_classes = [IsAuthenticated]
    http_method_names = ['delete']

    def delete(self, request):
        user = request.user
        try:
            with transaction.atomic():
                user.delete()
            return Response(
                {"detail": "Your account has been deleted successfully."},
                status=status.HTTP_204_NO_CONTENT
            )
        except IntegrityError as e:
            return Response(
                {"detail": f"Account deletion failed: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )
